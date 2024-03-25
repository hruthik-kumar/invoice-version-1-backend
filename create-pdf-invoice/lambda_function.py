import openpyxl
import psycopg2
import sql_query
import logging
import boto3
from excel_mapping import INVOICE_MAPPING
import traceback

logger = logging.getLogger('create-pdf-invoice')
logger.setLevel(0)

class ExecutionError(Exception):
    "Custome class to raise execution error."

def get_invoice_details_of_invoice_id(connect, invoice_id):
    try:
        cursor = connect.cursor()
        cursor.execute(sql_query.GET_INVOICE_DETAILS, (invoice_id,))
        result = cursor.fetchall()
        logger.debug("Invoice details: %s", result)
        if result:
            return result[0]
        raise ExecutionError("Empty invoice information was fetched.")
    except Exception as exc:
        logger.error("An error occured in fetching invoice details: %s", exc)
        raise ExecutionError("An Execution Error occured in finding invoice details.")

def download_invoice_template():
    try:
        s3 = boto3.client('s3')
        bucket_name = 'your_bucket_name'
        object_name = 'template.xlsx'
        local_file_path = 'temp/template.xlsx'
        s3.download_file(bucket_name, object_name, local_file_path)
        return local_file_path
    except Exception as exc:
        logger.error("An error occured in fetching invoice details: %s", exc)
        raise ExecutionError("An Execution Error occured in filling the invoice details.")

def edit_excel_file(file_path, invoice_details):
    """
    Edit an Excel file using openpyxl
    :param file_path: Path to the Excel file
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the active worksheet
    ws = wb.active
    
    # Modify content in the Excel sheet
    for key,val in INVOICE_MAPPING.items():
        for value in val:
            existing_content = ws[value].value
            existing_content = existing_content if existing_content is not None else '' 
            existing_content += str(invoice_details[key]) if invoice_details[key] is not None else ''
            ws[value] = existing_content
    
    # Save the workbook
    wb.save(file_path)
    print(f"Excel file '{file_path}' has been edited and saved successfully.")

def lambda_handler(event, contents):
    try:
        # invoice_id = event["query"]["invoice_id"]
        invoice_id = 2
        connect = psycopg2.connect(
            host= "localhost",
            dbname = "postgres",
            port= "5432",
            user= "postgres",
            password= "AdminSinga@6530",
        )
        invoice_details = get_invoice_details_of_invoice_id(connect, invoice_id)
        logger.debug('invoice_details : %s', invoice_details)
        
        # local_file_path = download_invoice_template()
        local_file_path = 'F:\dula_projects\Hruthik\Backend\invoice-version-1-backend\create-pdf-invoice\Invoice_Template2.xlsx'

        edit_excel_file(local_file_path, invoice_details)

        return {
            'statusCode' : 200,
            'body' : 'Invoice PDF generated successfully.'
        }

    except Exception as exc:
        traceback.print_exc()
        logger.error('Error occured : %s', exc)
        return {
            'statusCode' : 500,
            'body' : 'Error occured during execution.'
        }

res = lambda_handler(1,2)
print(res)